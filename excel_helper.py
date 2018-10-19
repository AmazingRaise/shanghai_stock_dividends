#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/7/4 下午9:11
# @Author  : Aries
# @Site    : 
# @File    : excel_helper.py
# @Software: PyCharm Community Edition
import logging
import openpyxl
import time
import os
import threading
logger = logging.getLogger(__name__)


class ExcelHelper(object):

    def __init__(self, filename=None):
        self.file_name = filename
        self.wb = None
        self.active_sheet = None
        self.rows = None
        self.cols = None
        self.table = None
        self.mutex = threading.Lock()

    def _set_filename(self):
        week = time.strftime("%W")
        year = time.strftime("%Y")
        file_name = year + '_' + week + '.xlsx'
        self.file_name = os.path.abspath(os.path.join('./data', file_name))

    def insert_data(self, content):
        """
        :param content: list类型，且为一层 
        :return: None
        """
        self.mutex.acquire()
        try:
            if not self.file_name:
                self._set_filename()
            if os.path.isfile(self.file_name):
                self.load()
            else:
                self.new_excel()
            self.append_table(content)
        except Exception as e:
            logger.info(e)
        finally:
            self.mutex.release()

    def load(self):
        self.wb = openpyxl.load_workbook(self.file_name)
        self.table = self.wb.active
        self.get_rows()

    def new_excel(self):
        self.wb = openpyxl.Workbook()
        self.table = self.wb.active
        self.rows = 0
        self.cols = 0

    def append_table(self, content):
        if isinstance(content, list):
            # 追加到最后
            for val in content:
                if isinstance(val, list):
                    self.table.append(val)
                else:
                    self.table.append(content)
                    break

    def save_excel(self):
        self.wb.save(self.file_name)

    def get_rows(self):
        self.rows = self.table.max_row   #获取行数
        self.cols = self.table.max_column    #获取列数

    def get_sheets(self):
        return self.wb.sheet_names


class ExcelReader(object):

    def __init__(self):
        self.filename = None
        self.wb = None
        self.sheet = None
        self.result = []

    def start(self, filename):
        self.load_excel(filename)
        return self.get_content()

    def load_excel(self, filename):
        self.wb = openpyxl.load_workbook(filename)
        self.sheet = self.wb.active

    def get_content(self):
        for row in self.sheet.rows:
            row_li = []
            for cell in row:
                row_li.append(cell.value)
            self.result.append(row_li)
        return self.result

